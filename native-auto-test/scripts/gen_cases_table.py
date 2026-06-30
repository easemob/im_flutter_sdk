#!/usr/bin/env python3
"""Extract all pytest test cases and generate a Chinese cases table (markdown).

Description priority:
  1) per-case record docs: docs/agents/<module>/CASES_RECORD.zh.md
     (provides API grouping, normal/exception category, and rich description)
  2) the test function docstring
  3) humanized function name fallback
"""
from __future__ import annotations

import ast
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TESTS_DIR = ROOT / "tests"
DOCS_DIR = ROOT / "docs" / "agents"
OUT = ROOT / "docs" / "CASES_TABLE.zh.md"
OUT_XLSX = ROOT / "docs" / "CASES_TABLE.zh.xlsx"

HEADER = ["API所属模块", "API覆盖项", "用例函数", "功能描述", "操作步骤", "预期结果", "测试文件", "执行", "级别", "备注"]

MODULE_NAMES = {
    "chat": "Chat 消息",
    "chatroom": "ChatRoom 聊天室",
    "client": "Client 客户端",
    "contact": "Contact 好友",
    "group": "Group 群组",
    "presence": "Presence 在线状态",
    "push": "Push 推送",
    "user_info": "UserInfo 用户属性",
}

CASE_REF_RE = re.compile(r"`(tests/[^`]+?\.py)::(test_[A-Za-z0-9_]+)`")


def parse_record_docs():
    """Return map: func_name -> {api, category, desc, file}."""
    info = {}
    for rec in sorted(DOCS_DIR.rglob("CASES_RECORD.zh.md")):
        cur_api = ""
        cur_cat = ""
        lines = rec.read_text(encoding="utf-8").splitlines()
        i = 0
        while i < len(lines):
            line = lines[i]
            s = line.strip()
            if s.startswith("## "):
                cur_api = s[3:].strip()
                cur_cat = ""
            elif s.startswith("正常 cases") or s == "正常 cases":
                cur_cat = "正常"
            elif s.startswith("异常 cases") or s == "异常 cases":
                cur_cat = "异常"
            m = CASE_REF_RE.search(line)
            if m:
                fpath, func = m.group(1), m.group(2)
                # description: following indented non-empty lines
                desc_parts = []
                j = i + 1
                while j < len(lines):
                    nxt = lines[j]
                    if not nxt.strip():
                        break
                    if CASE_REF_RE.search(nxt) or nxt.strip().startswith("##") \
                            or nxt.strip().startswith("正常 cases") \
                            or nxt.strip().startswith("异常 cases"):
                        break
                    # numbered new item line
                    if re.match(r"^\s*\d+\.\s", nxt):
                        break
                    desc_parts.append(nxt.strip())
                    j += 1
                desc = " ".join(desc_parts)
                # keep first occurrence (don't overwrite with link-only later refs)
                if func not in info or (not info[func]["desc"] and desc):
                    info[func] = {
                        "api": cur_api,
                        "category": cur_cat,
                        "desc": desc,
                        "file": fpath,
                    }
                i = j
                continue
            i += 1
    return info


def docstring_api_desc(doc: str):
    doc = re.sub(r"\s+", " ", (doc or "").strip())
    if not doc:
        return "", ""
    m = re.split(r"[：:]", doc, maxsplit=1)
    if len(m) == 2 and len(m[0]) <= 60 and " " not in m[0].strip():
        return m[0].strip(), m[1].strip()
    return "", doc


def cell(s: str) -> str:
    return (s or "").replace("|", "\\|").replace("\n", " ").strip()


EXPECT_KEYWORDS = (
    "验证", "校验", "断言", "冻结", "预期", "确保", "确认", "返回", "收到",
    "成功", "失败", "错误", "报错", "不应", "应", "幂等", "可见", "不可见",
    "一致", "稳定", "完整", "正确", "保留", "清空", "清零", "命中", "回传",
    "回调", "为空", "非空", "抛出", "拒绝", "生效", "更新", "同步",
)


def expected_from_desc(desc: str, category: str) -> str:
    """Extract the expectation clauses from a Chinese case description."""
    desc = re.sub(r"\s+", " ", (desc or "").strip())
    if not desc:
        return "异常入参返回稳定的错误语义" if category == "异常" else "操作成功且返回字段符合预期"
    segs = re.split(r"[；;。，,]", desc)
    segs = [s.strip() for s in segs if s.strip()]
    picked = [s for s in segs if any(k in s for k in EXPECT_KEYWORDS)]
    if picked:
        out = "；".join(picked)
    else:
        out = segs[-1] if segs else desc
    out = re.sub(r"^(再|然后|并|且|最后|随后)\s*", "", out)
    return out


def build_rows():
    """Return (rows, counts) for all test_ functions across tests/."""
    rec = parse_record_docs()
    rows = []
    counts = {}
    for py in sorted(TESTS_DIR.rglob("test_*.py")):
        rel = str(py.relative_to(TESTS_DIR.parent))
        parts = py.relative_to(TESTS_DIR).parts
        module = parts[0] if len(parts) > 1 else "root"
        tree = ast.parse(py.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)) and node.name.startswith("test_"):
                doc = ast.get_docstring(node) or ""
                d_api, d_desc = docstring_api_desc(doc)
                r = rec.get(node.name)
                if r and r["desc"]:
                    api = r["api"] or d_api
                    desc = r["desc"]
                    category = r["category"]
                elif d_desc:
                    api, desc, category = d_api, d_desc, ""
                else:
                    api, category = d_api, ""
                    desc = node.name.replace("test_", "").replace("_", " ")
                rows.append({
                    "module": module,
                    "api": api,
                    "func": node.name,
                    "desc": desc,
                    "category": category,
                    "expected": expected_from_desc(desc, category),
                    "file": rel.replace("native-auto-test/", ""),
                })
                counts[module] = counts.get(module, 0) + 1
    rows.sort(key=lambda r: (r["module"], r["file"], r["func"]))
    return rows, counts


def main():
    rows, counts = build_rows()
    matched_rec = sum(1 for r in rows if r["category"])
    matched_doc = 0

    L = []
    L.append("# 自动化测试用例总表（native-auto-test）\n")
    L.append(f"> 自动生成（含未提交改动）。统计口径：`tests/**/test_*.py` 中的 `test_` 函数，共 **{len(rows)}** 条。\n")
    L.append("## 用例数量统计（按模块）\n")
    L.append("| API 所属模块 | 用例数 |")
    L.append("|---|---|")
    for m in sorted(counts):
        L.append(f"| {MODULE_NAMES.get(m, m)} | {counts[m]} |")
    L.append(f"| **合计** | **{len(rows)}** |")
    L.append("")
    L.append("## 用例明细表\n")
    header = HEADER
    L.append("| " + " | ".join(header) + " |")
    L.append("|" + "|".join(["---"] * len(header)) + "|")
    for r in rows:
        L.append("| " + " | ".join([
            cell(MODULE_NAMES.get(r["module"], r["module"])),
            cell(r["api"]),
            cell(r["func"]),
            cell(r["desc"]),
            "",  # 操作步骤（人工补充）
            "",  # 预期结果（人工补充）
            cell(r["file"]),
            "",  # 执行（人工/CI 回填）
            cell(r["category"]),  # 级别：正常/异常
            "",  # 备注
        ]) + " |")
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(L) + "\n", encoding="utf-8")
    print(f"wrote {OUT}")
    print(f"total={len(rows)} with_category={matched_rec}")
    write_xlsx(rows, counts)


def write_xlsx(rows, counts):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed; skip xlsx. run: pip install openpyxl")
        return

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Detail sheet (first / active: every case, like the md) ----
    ws = wb.active
    ws.title = "用例明细"
    ws.append(HEADER)
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for r in rows:
        ws.append([
            MODULE_NAMES.get(r["module"], r["module"]),
            r["api"],
            r["func"],
            r["desc"],
            "",  # 操作步骤
            "",  # 预期结果
            r["file"],
            "",  # 执行
            r["category"],  # 级别
            "",  # 备注
        ])
    widths = [16, 28, 42, 60, 24, 24, 40, 8, 8, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wrap_cols = {2, 4, 5, 6}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEADER)):
        for c in row:
            c.border = border
            wrap = c.column in wrap_cols
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}{ws.max_row}"

    # ---- Summary sheet (second) ----
    ws0 = wb.create_sheet("统计")
    ws0.append(["API 所属模块", "用例数"])
    for c in ws0[1]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    for m in sorted(counts):
        ws0.append([MODULE_NAMES.get(m, m), counts[m]])
    ws0.append(["合计", sum(counts.values())])
    total_row = ws0.max_row
    for c in ws0[total_row]:
        c.font = Font(bold=True)
    for row in ws0.iter_rows(min_row=2, max_row=ws0.max_row, max_col=2):
        for c in row:
            c.border = border
            c.alignment = Alignment(horizontal="left", vertical="center")
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 10

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
