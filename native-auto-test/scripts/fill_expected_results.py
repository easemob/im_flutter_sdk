#!/usr/bin/env python3
"""Fill the 预期结果 column of an external cases xlsx, matching by 用例函数.

Usage: python3 scripts/fill_expected_results.py /path/to/file.xlsx
Derives expected results from the same data source as gen_cases_table.py.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment

# reuse build logic
sys.path.insert(0, str(Path(__file__).resolve().parent))
from gen_cases_table import build_rows  # noqa: E402

DEFAULT = "/Users/kk/Desktop/隔舱测试/im-flutter-sdk.xlsx"


def main():
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT)
    if not target.exists():
        print(f"file not found: {target}")
        sys.exit(1)

    rows, _ = build_rows()
    by_func = {r["func"]: r for r in rows}

    wb = load_workbook(target)
    if "用例明细" in wb.sheetnames:
        ws = wb["用例明细"]
    else:
        ws = wb[wb.sheetnames[0]]

    header = [ (c.value or "").strip() if isinstance(c.value, str) else c.value for c in ws[1] ]
    # locate columns by header name
    def col_idx(name):
        for i, h in enumerate(header, start=1):
            if h == name:
                return i
        return None

    func_col = col_idx("用例函数")
    exp_col = col_idx("预期结果")
    if func_col is None or exp_col is None:
        print(f"missing columns. header={header}")
        sys.exit(1)

    filled = 0
    missed = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        func = row[func_col - 1].value
        if not func:
            continue
        rec = by_func.get(str(func).strip())
        if not rec:
            missed.append(func)
            continue
        target_cell = row[exp_col - 1]
        if target_cell.value:  # don't overwrite manual edits
            continue
        target_cell.value = rec["expected"]
        target_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        filled += 1

    wb.save(target)
    print(f"saved {target}; filled={filled}; unmatched={len(missed)}")
    if missed:
        print("unmatched funcs:", missed[:20])


if __name__ == "__main__":
    main()
